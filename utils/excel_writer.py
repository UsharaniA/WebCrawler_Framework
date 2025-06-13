import os
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

load_dotenv()

def write_to_excel(sheet_name, option_texts, file_path):
    try:
        max_versions = int(os.getenv("MAX_SHEET_VERSIONS", 99))
        max_excel_sheet_len = 31  # Excel's fixed limit

        # Load or create workbook
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)

        # Prepare base name and initialize
        base_name = sheet_name[:max_excel_sheet_len]
        sheet_name = base_name
        version = 1

        # Versioning logic
        while sheet_name in wb.sheetnames:
            suffix = f"_v{version}"
            base_truncated = base_name[:max_excel_sheet_len - len(suffix)]
            sheet_name = f"{base_truncated}{suffix}"
            version += 1

            if version > max_versions:
                raise ValueError(f"Exceeded max allowed sheet versions ({max_versions}) for '{base_name}'.")

        # Write data
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Index", "Option Text", "Estimated Time"])
        for idx, (option, estimated) in enumerate(option_texts, start=1):
            ws.append([idx, option, estimated])

        wb.save(file_path)
        print(f"[SUCCESS] Data written to sheet '{sheet_name}' in {file_path}")

    except Exception as e:
        print(f"[ERROR] Writing to Excel failed: {e}")
